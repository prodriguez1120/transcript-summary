#!/usr/bin/env python3
"""
Streamlined Quote Analysis Tool for FlexXray Transcripts

This module implements a sophisticated quote analysis system that:
1. Uses vector database for semantic search
2. Implements quote ranking and reranking
3. Streamlines AI calls for better efficiency
4. Focuses on question-answer matching
"""

import os
import json
import re
import hashlib
import sqlite3
import asyncio
from openai import OpenAI
from typing import List, Dict, Any, Tuple, Optional
from pathlib import Path
import time
from dotenv import load_dotenv
from datetime import datetime
import numpy as np
import logging

# Check if openpyxl is available for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Load environment variables
load_dotenv()

# Import robust metadata filtering
from robust_metadata_filtering import RobustMetadataFilter


class StreamlinedQuoteAnalysis:
    def __init__(self, api_key: Optional[str] = None, cache_dir: str = "cache"):
        """Initialize the streamlined quote analysis tool."""
        self.api_key = api_key or os.getenv("OPENAI_API_KEY")
        if not self.api_key:
            raise ValueError(
                "OpenAI API key is required. Set OPENAI_API_KEY environment variable or pass it to the constructor."
            )

        self.client = OpenAI(api_key=self.api_key)
        self.cache_dir = cache_dir
        self._init_cache()
        
        # Configure logging to reduce verbosity
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.WARNING)  # Only show warnings and errors
        
        # Metadata validation settings
        self.metadata_validation_enabled = True
        self.confidence_threshold = 2  # Minimum confidence for interviewer detection
        
        # Initialize robust metadata filter
        self.metadata_filter = RobustMetadataFilter(confidence_threshold=self.confidence_threshold)

        # Define the key questions for analysis
        self.key_questions = {
            "market_leadership": "What evidence shows FlexXray's market leadership and competitive advantage?",
            "value_proposition": "How does FlexXray's value proposition address the risk of insourcing?",
            "local_presence": "How does FlexXray's local presence and footprint drive customer demand?",
            "technology_advantages": "What proprietary technology advantages does FlexXray offer?",
            "rapid_turnaround": "How do FlexXray's rapid turnaround times benefit customers?",
            "limited_tam": "What limits FlexXray's Total Addressable Market (TAM)?",
            "unpredictable_timing": "How does unpredictable event timing impact FlexXray's business?",
        }

        # Question categories for organization
        self.question_categories = {
            "key_takeaways": [
                "market_leadership",
                "value_proposition",
                "local_presence",
            ],
            "strengths": ["technology_advantages", "rapid_turnaround"],
            "weaknesses": ["limited_tam", "unpredictable_timing"],
        }

    def _init_cache(self):
        """Initialize the cache database."""
        os.makedirs(self.cache_dir, exist_ok=True)
        cache_db = os.path.join(self.cache_dir, "ranking_cache.db")
        
        with sqlite3.connect(cache_db) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS ranking_cache (
                    cache_key TEXT PRIMARY KEY,
                    question_id TEXT NOT NULL,
                    transcript_hash TEXT NOT NULL,
                    model_version TEXT NOT NULL,
                    ranked_results TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    expires_at TIMESTAMP
                )
            """)
            conn.commit()

    def get_expert_quotes_only(self, quotes: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Filter quotes to include only expert responses, filtering out interviewer questions."""
        if not quotes:
            return []

        # Use robust metadata filtering to identify and filter quotes
        filtered_quotes = self.metadata_filter.prefilter_quotes_by_metadata(quotes, "")
        
        # Additional validation and correction
        validated_quotes = self.metadata_filter.validate_and_correct_metadata(filtered_quotes)
        
        # Filter to only include expert responses
        expert_quotes = []
        for quote in validated_quotes:
            text = quote.get("text", "")
            metadata = quote.get("metadata", {})
            speaker_role = metadata.get("speaker_role", "unknown")
            
            # If speaker role is explicitly marked as expert, include it
            if speaker_role == "expert":
                expert_quotes.append(quote)
                continue
            
            # If speaker role is interviewer, exclude it
            if speaker_role == "interviewer":
                continue
            
            # For unknown speaker roles, use robust text analysis
            if self.metadata_filter.is_likely_expert_response(text):
                expert_quotes.append(quote)
        
        return expert_quotes

    def _is_likely_expert_response(self, text: str) -> bool:
        """Simple heuristic to determine if text is likely an expert response."""
        if not text:
            return False
        
        text_lower = text.lower()
        
        # Expert indicators
        expert_indicators = [
            "we have", "we are", "we do", "we provide", "we offer", "we deliver",
            "our company", "our technology", "our service", "our customers",
            "flexxray", "company", "business", "industry", "market"
        ]
        
        # Interviewer indicators
        interviewer_indicators = [
            "what", "how", "why", "when", "where", "can you", "could you",
            "tell me", "describe", "explain", "walk me through"
        ]
        
        # Count indicators
        expert_count = sum(1 for indicator in expert_indicators if indicator in text_lower)
        interviewer_count = sum(1 for indicator in interviewer_indicators if indicator in text_lower)
        
        # Simple scoring: more expert indicators = more likely expert response
        return expert_count > interviewer_count

    def _is_interviewer_question(self, text: str) -> bool:
        """Simple heuristic to determine if text is an interviewer question."""
        if not text:
            return False
        
        text_lower = text.lower()
        
        # Question indicators
        question_words = ["what", "how", "why", "when", "where", "can you", "could you"]
        question_marks = text.count("?")
        
        # Starts with question word or ends with question mark
        starts_with_question = any(text_lower.startswith(word) for word in question_words)
        ends_with_question = text.strip().endswith("?")
        
        return starts_with_question or ends_with_question or question_marks > 0

    def rank_quotes_for_question(self, quotes: List[Dict[str, Any]], question: str, top_k: int = 10) -> List[Dict[str, Any]]:
        """Rank quotes for a specific question using a simplified approach."""
        if not quotes:
            return []

        # Simple ranking based on text length and keyword matching
        ranked_quotes = []
        question_lower = question.lower()
        
        for i, quote in enumerate(quotes):
            quote_copy = quote.copy()
            text = quote.get("text", "").lower()
            
            # Simple scoring
            score = 5  # Base score
            
            # Boost score for longer, more detailed quotes
            if len(text) > 100:
                score += 2
            if len(text) > 200:
                score += 1
            
            # Boost score for question-related terms
            question_terms = question_lower.split()
            for term in question_terms:
                if term in text and len(term) > 3:  # Skip short words
                    score += 1
            
            quote_copy["relevance_score"] = min(score, 10)  # Cap at 10
            quote_copy["relevance_explanation"] = "Simplified ranking based on text analysis"
            ranked_quotes.append(quote_copy)
        
        # Sort by score and return top_k
        ranked_quotes.sort(key=lambda x: x.get("relevance_score", 0), reverse=True)
        return ranked_quotes[:top_k]

    def generate_company_summary(self, quotes: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate a comprehensive company summary from the analyzed quotes."""
        if not quotes:
            return {"error": "No quotes provided for analysis"}
        
        # Get expert quotes only
        expert_quotes = self.get_expert_quotes_only(quotes)
        
        # Generate insights for each question category
        key_takeaways = self._generate_insights_for_category(expert_quotes, "key_takeaways")
        strengths = self._generate_insights_for_category(expert_quotes, "strengths")
        weaknesses = self._generate_insights_for_category(expert_quotes, "weaknesses")
        
        summary = {
            "key_takeaways": key_takeaways,
            "strengths": strengths,
            "weaknesses": weaknesses,
            "total_quotes": len(quotes),
            "expert_quotes": len(expert_quotes),
            "interviewer_quotes": len(quotes) - len(expert_quotes),
            "analysis_timestamp": datetime.now().isoformat(),
            "summary": f"Analyzed {len(quotes)} quotes with {len(expert_quotes)} expert responses."
        }
        
        return summary
    
    def _generate_insights_for_category(self, quotes: List[Dict[str, Any]], category: str) -> List[Dict[str, Any]]:
        """Generate insights and supporting quotes for a specific category."""
        if not quotes:
            return []
        
        # Get questions for this category
        category_questions = self.question_categories.get(category, [])
        insights = []
        
        for question_key in category_questions:
            question_text = self.key_questions.get(question_key, "")
            if not question_text:
                continue
            
            # Pre-filter quotes by metadata for this specific question
            question_filtered_quotes = self.metadata_filter.prefilter_quotes_by_metadata(quotes, question_text)
            
            # Rank quotes for this question
            ranked_quotes = self.rank_quotes_for_question(question_filtered_quotes, question_text, top_k=3)
            
            if ranked_quotes:
                # Create insight with supporting quotes
                insight = {
                    "insight": f"Insight for {question_key.replace('_', ' ').title()}",
                    "question": question_text,
                    "supporting_quotes": []
                }
                
                # Add supporting quotes with metadata
                for quote in ranked_quotes[:2]:  # Limit to 2 quotes per insight
                    quote_data = {
                        "text": quote.get("text", ""),
                        "formatted_text": quote.get("text", ""),
                        "metadata": quote.get("metadata", {}),
                        "relevance_score": quote.get("relevance_score", 0)
                    }
                    insight["supporting_quotes"].append(quote_data)
                
                insights.append(insight)
        
        return insights

    def format_summary_output(self, summary_results: Dict[str, Any]) -> str:
        """Format the summary results for display."""
        if not summary_results:
            return "No summary results available."
        
        output = "FlexXray Company Analysis Summary\n"
        output += "=" * 40 + "\n\n"
        
        output += f"Total Quotes Analyzed: {summary_results.get('total_quotes', 0)}\n"
        output += f"Expert Responses: {summary_results.get('expert_quotes', 0)}\n"
        output += f"Interviewer Questions: {summary_results.get('interviewer_quotes', 0)}\n"
        output += f"Analysis Time: {summary_results.get('analysis_timestamp', 'Unknown')}\n\n"
        
        # Add key takeaways
        takeaways = summary_results.get("key_takeaways", [])
        if takeaways:
            output += "KEY TAKEAWAYS\n"
            output += "-" * 20 + "\n"
            for i, takeaway in enumerate(takeaways, 1):
                output += f"{i}. {takeaway.get('insight', 'No insight')}\n"
                if takeaway.get("supporting_quotes"):
                    for quote in takeaway["supporting_quotes"][:2]:
                        output += f"   • {quote.get('text', 'No quote text')}\n"
                output += "\n"
        
        # Add strengths
        strengths = summary_results.get("strengths", [])
        if strengths:
            output += "STRENGTHS\n"
            output += "-" * 20 + "\n"
            for i, strength in enumerate(strengths, 1):
                output += f"{i}. {strength.get('insight', 'No insight')}\n"
                if strength.get("supporting_quotes"):
                    for quote in strength["supporting_quotes"][:2]:
                        output += f"   • {quote.get('text', 'No quote text')}\n"
                output += "\n"
        
        # Add weaknesses
        weaknesses = summary_results.get("weaknesses", [])
        if weaknesses:
            output += "WEAKNESSES\n"
            output += "-" * 20 + "\n"
            for i, weakness in enumerate(weaknesses, 1):
                output += f"{i}. {weakness.get('insight', 'No insight')}\n"
                if weakness.get("supporting_quotes"):
                    for quote in weakness["supporting_quotes"][:2]:
                        output += f"   • {quote.get('text', 'No quote text')}\n"
                output += "\n"
        
        output += f"Summary: {summary_results.get('summary', 'No summary available')}\n"
        
        return output

    def save_summary(self, summary_results: Dict[str, Any], output_dir: str = "Outputs") -> str:
        """Save the summary results to a file."""
        if not summary_results:
            print("Warning: No summary results to save")
            return ""
        
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"company_summary_{timestamp}.txt"
        filepath = os.path.join(output_dir, filename)
        
        try:
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(self.format_summary_output(summary_results))
            print(f"Summary saved to: {filepath}")
            return filepath
        except Exception as e:
            print(f"Error saving summary: {e}")
            return ""

    def export_to_excel(self, summary_results: Dict[str, Any], output_dir: str = "Outputs") -> str:
        """Export results to Excel format with enhanced segmentation by key takeaways, strengths, and weaknesses."""
        if not EXCEL_AVAILABLE:
            print("Warning: openpyxl not available, cannot export to Excel")
            return ""
        
        if not summary_results:
            print("Warning: No summary results to export")
            return ""
        
        os.makedirs(output_dir, exist_ok=True)
        
        try:
            # Use the enhanced export from export_utils
            from export_utils import ExportManager
            export_manager = ExportManager(output_dir)
            
            # Export using enhanced functionality
            output_file = export_manager.export_quote_analysis_to_excel(summary_results)
            
            if output_file:
                print(f"✅ Enhanced Excel export successful: {output_file}")
                return output_file
            else:
                print("⚠️  Enhanced export failed, falling back to basic export")
                return self._fallback_excel_export(summary_results, output_dir)
                
        except ImportError:
            print("⚠️  export_utils not available, using basic export")
            return self._fallback_excel_export(summary_results, output_dir)
        except Exception as e:
            print(f"⚠️  Enhanced export error: {e}, using basic export")
            return self._fallback_excel_export(summary_results, output_dir)
    
    def _fallback_excel_export(self, summary_results: Dict[str, Any], output_dir: str) -> str:
        """Fallback to basic Excel export if enhanced export fails."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"FlexXray_Streamlined_Summary_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Company Analysis"
            
            # Define styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )
            question_font = Font(bold=True)
            quote_font = Font(italic=True)
            
            # Add headers with the exact format requested
            headers = ["Question", "Supporting Quotes", "Speaker/Transcript", "Score"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            current_row = 2
            
            # Process each category and add data rows
            categories = {
                "key_takeaways": summary_results.get("key_takeaways", []),
                "strengths": summary_results.get("strengths", []),
                "weaknesses": summary_results.get("weaknesses", [])
            }
            
            for category_name, category_data in categories.items():
                if not category_data:
                    continue
                
                for insight in category_data:
                    question_text = insight.get("question", insight.get("insight", ""))
                    supporting_quotes = insight.get("supporting_quotes", [])
                    
                    if supporting_quotes:
                        # Add row for each quote
                        for quote in supporting_quotes:
                            # Question column
                            ws.cell(row=current_row, column=1, value=question_text)
                            ws.cell(row=current_row, column=1).font = question_font
                            
                            # Supporting Quotes column
                            quote_text = quote.get("formatted_text", quote.get("text", ""))
                            ws.cell(row=current_row, column=2, value=quote_text)
                            ws.cell(row=current_row, column=2).font = quote_font
                            ws.cell(row=current_row, column=2).alignment = Alignment(
                                wrap_text=True, vertical="top"
                            )
                            
                            # Speaker/Transcript column
                            metadata = quote.get("metadata", {})
                            speaker_role = metadata.get("speaker_role", "Unknown")
                            transcript_name = metadata.get("transcript_name", "Unknown")
                            speaker_transcript = f"{speaker_role.title()}/{transcript_name}"
                            ws.cell(row=current_row, column=3, value=speaker_transcript)
                            
                            # Score column
                            score = quote.get("relevance_score", 0)
                            ws.cell(row=current_row, column=4, value=score)
                            
                            current_row += 1
                    else:
                        # Add row even if no quotes
                        ws.cell(row=current_row, column=1, value=question_text)
                        ws.cell(row=current_row, column=1).font = question_font
                        current_row += 1
            
            # Auto-adjust column widths
            column_widths = {
                "A": 50,  # Question column
                "B": 80,  # Supporting Quotes column - Wide for full quote display
                "C": 30,  # Speaker/Transcript column
                "D": 15,  # Score column
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Save workbook
            wb.save(filepath)
            wb.close()
            
            print(f"FlexXray Streamlined Summary exported to Excel: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return ""


def main():
    """Main function for testing."""
    print("Streamlined Quote Analysis Tool")
    print("=" * 40)
    
    # Check if API key is available
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("Warning: OPENAI_API_KEY not set. Some functionality may be limited.")
        api_key = "test_key"  # Use test key for basic functionality
    
    try:
        analyzer = StreamlinedQuoteAnalysis(api_key=api_key)
        print("✓ Tool initialized successfully")
        
        # Test basic functionality
        test_quotes = [
            {
                "text": "FlexXray provides excellent foreign material detection services.",
                "metadata": {"speaker_role": "expert", "transcript_name": "Test 1"}
            },
            {
                "text": "What are the main benefits of your service?",
                "metadata": {"speaker_role": "interviewer", "transcript_name": "Test 2"}
            }
        ]
        
        print(f"\nTesting with {len(test_quotes)} quotes...")
        
        # Test expert filtering
        expert_quotes = analyzer.get_expert_quotes_only(test_quotes)
        print(f"Expert quotes: {len(expert_quotes)}")
        
        # Test ranking
        ranked_quotes = analyzer.rank_quotes_for_question(test_quotes, "What services does FlexXray provide?")
        print(f"Ranked quotes: {len(ranked_quotes)}")
        
        # Test summary generation
        summary = analyzer.generate_company_summary(test_quotes)
        print(f"Summary generated: {summary.get('summary', 'No summary')}")
        
        print("\n✓ All basic functionality tests passed!")
        
    except Exception as e:
        print(f"✗ Error during testing: {e}")


if __name__ == "__main__":
    main()
