#!/usr/bin/env python3
"""
Export Utilities Module for FlexXray Transcripts

This module handles export functionality including Excel, text, and other formats
for the quote analysis tool.
"""

import os
import json
import re
from typing import List, Dict, Any, Optional
from datetime import datetime

# Check if openpyxl is available for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExportManager:
    def __init__(self, output_directory: str = "Outputs"):
        """Initialize the export manager."""
        self.output_directory = output_directory
        os.makedirs(output_directory, exist_ok=True)

    def save_quote_analysis(
        self, results: Dict[str, Any], output_file: str = None
    ) -> str:
        """Save quote analysis results to JSON file."""
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(
                self.output_directory, f"FlexXray_quote_analysis_{timestamp}.json"
            )

        try:
            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(results, f, indent=2, ensure_ascii=False)

            print(f"Quote analysis saved to: {output_file}")
            return output_file

        except (IOError, OSError) as e:
            print(f"File system error saving quote analysis: {e}")
            return ""
        except (ValueError, TypeError) as e:
            print(f"Data serialization error: {e}")
            return ""
        except Exception as e:
            print(f"Unexpected error saving quote analysis: {e}")
            return ""

    def export_quote_analysis_to_text(
        self, results: Dict[str, Any], output_file: str = None
    ) -> str:
        """Export quote analysis results to text file."""
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(
                self.output_directory, f"FlexXray_quote_analysis_{timestamp}.txt"
            )

        try:
            with open(output_file, "w", encoding="utf-8") as f:
                f.write("FLEXXRAY QUOTE ANALYSIS RESULTS\n")
                f.write("=" * 50 + "\n\n")

                # Write metadata
                metadata = results.get("metadata", {})
                f.write(f"Analysis Date: {metadata.get('analysis_date', 'Unknown')}\n")
                f.write(f"Total Transcripts: {metadata.get('total_transcripts', 0)}\n")
                f.write(f"Total Quotes: {metadata.get('total_quotes', 0)}\n\n")

                # Write perspectives
                perspectives = results.get("perspectives", {})
                for perspective_key, perspective_data in perspectives.items():
                    f.write(
                        f"PERSPECTIVE: {perspective_data.get('title', 'Unknown')}\n"
                    )
                    f.write("-" * 40 + "\n")
                    f.write(
                        f"Description: {perspective_data.get('description', '')}\n\n"
                    )

                    # Write themes
                    themes = perspective_data.get("themes", [])
                    for theme in themes:
                        f.write(f"  Theme: {theme.get('name', 'Unknown')}\n")
                        f.write(f"  Description: {theme.get('description', '')}\n")

                        # Write quotes for this theme
                        quotes = theme.get("quotes", [])
                        for i, quote in enumerate(quotes, 1):
                            f.write(f"    Quote {i}: {quote.get('text', '')}\n")
                            if quote.get("transcript_name"):
                                f.write(f"      Source: {quote['transcript_name']}\n")
                        f.write("\n")

                    f.write("\n")

                # Write quote summary
                quote_summary = results.get("quote_summary", {})
                f.write("QUOTE SUMMARY\n")
                f.write("-" * 20 + "\n")
                f.write(f"Strengths: {quote_summary.get('strengths_count', 0)}\n")
                f.write(f"Weaknesses: {quote_summary.get('weaknesses_count', 0)}\n")
                f.write(f"Neutral: {quote_summary.get('neutral_count', 0)}\n\n")

            print(f"Quote analysis exported to text: {output_file}")
            return output_file

        except (IOError, OSError) as e:
            print(f"File system error exporting to text: {e}")
            return ""
        except (ValueError, TypeError) as e:
            print(f"Data formatting error: {e}")
            return ""
        except Exception as e:
            print(f"Unexpected error exporting to text: {e}")
            return ""

    def export_company_summary_page(
        self, summary_data: Dict[str, Any], output_file: str = None
    ) -> str:
        """Export company summary page to text file."""
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(
                self.output_directory, f"FlexXray_Company_Summary_Page_{timestamp}.txt"
            )

        try:
            with open(output_file, "w", encoding="utf-8") as f:
                f.write("FLEXXRAY COMPANY SUMMARY PAGE\n")
                f.write("=" * 40 + "\n\n")

                # Write key takeaways
                takeaways = summary_data.get("key_takeaways", [])
                if takeaways:
                    f.write("KEY TAKEAWAYS\n")
                    f.write("-" * 20 + "\n")
                    for i, takeaway in enumerate(takeaways, 1):
                        # Handle both 'insight' and 'theme' fields
                        insight = takeaway.get("insight", takeaway.get("theme", ""))
                        f.write(f"{i}. {insight}\n")

                        # Handle both 'supporting_quotes' and 'quotes' fields
                        quotes = takeaway.get(
                            "supporting_quotes", takeaway.get("quotes", [])
                        )
                        if quotes:
                            f.write("   Supporting quotes:\n")
                            for quote in quotes:  # Show all quotes
                                if isinstance(quote, dict):
                                    # Use formatted_text if available, otherwise fall back to text
                                    quote_text = quote.get(
                                        "formatted_text",
                                        quote.get("text", quote.get("quote", "")),
                                    )
                                    speaker = quote.get(
                                        "speaker", quote.get("speaker_info", "Unknown")
                                    )
                                    document = quote.get(
                                        "document",
                                        quote.get("transcript_name", "Unknown"),
                                    )
                                    f.write(
                                        f'     - "{quote_text}" - {speaker} from {document}\n'
                                    )
                                else:
                                    f.write(f"     - {quote}\n")
                        f.write("\n")

                # Write strengths
                strengths = summary_data.get("strengths", [])
                if strengths:
                    f.write("STRENGTHS\n")
                    f.write("-" * 15 + "\n")
                    for i, strength in enumerate(strengths, 1):
                        # Handle both 'insight' and 'theme' fields
                        insight = strength.get("insight", strength.get("theme", ""))
                        f.write(f"{i}. {insight}\n")

                        # Handle both 'supporting_quotes' and 'quotes' fields
                        quotes = strength.get(
                            "supporting_quotes", strength.get("quotes", [])
                        )
                        if quotes:
                            f.write("   Supporting quotes:\n")
                            for quote in quotes:  # Show all quotes
                                if isinstance(quote, dict):
                                    # Use formatted_text if available, otherwise fall back to text
                                    quote_text = quote.get(
                                        "formatted_text",
                                        quote.get("text", quote.get("quote", "")),
                                    )
                                    speaker = quote.get(
                                        "speaker", quote.get("speaker_info", "Unknown")
                                    )
                                    document = quote.get(
                                        "document",
                                        quote.get("transcript_name", "Unknown"),
                                    )
                                    f.write(
                                        f'     - "{quote_text}" - {speaker} from {document}\n'
                                    )
                                else:
                                    f.write(f"     - {quote}\n")
                        f.write("\n")

                # Write weaknesses
                weaknesses = summary_data.get("weaknesses", [])
                if weaknesses:
                    f.write("WEAKNESSES\n")
                    f.write("-" * 15 + "\n")
                    for i, weakness in enumerate(weaknesses, 1):
                        # Handle both 'insight' and 'theme' fields
                        insight = weakness.get("insight", weakness.get("theme", ""))
                        f.write(f"{i}. {insight}\n")

                        # Handle both 'supporting_quotes' and 'quotes' fields
                        quotes = weakness.get(
                            "supporting_quotes", weakness.get("quotes", [])
                        )
                        if quotes:
                            f.write("   Supporting quotes:\n")
                            for quote in quotes:  # Show all quotes
                                if isinstance(quote, dict):
                                    # Use formatted_text if available, otherwise fall back to text
                                    quote_text = quote.get(
                                        "formatted_text",
                                        quote.get("text", quote.get("quote", "")),
                                    )
                                    speaker = quote.get(
                                        "speaker", quote.get("speaker_info", "Unknown")
                                    )
                                    document = quote.get(
                                        "document",
                                        quote.get("transcript_name", "Unknown"),
                                    )
                                    f.write(
                                        f'     - "{quote_text}" - {speaker} from {document}\n'
                                    )
                                else:
                                    f.write(f"     - {quote}\n")
                        f.write("\n")

            print(f"Company summary page exported to text: {output_file}")
            return output_file

        except (IOError, OSError) as e:
            print(f"File system error exporting company summary page: {e}")
            return ""
        except (ValueError, TypeError) as e:
            print(f"Data formatting error exporting company summary page: {e}")
            return ""
        except Exception as e:
            print(f"Unexpected error exporting company summary page: {e}")
            return ""

    def export_company_summary_to_excel(
        self, summary_data: Dict[str, Any], output_file: str = None
    ) -> str:
        """Export company summary page to Excel file."""
        if not EXCEL_AVAILABLE:
            print("openpyxl not available for Excel export")
            return ""

        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(
                self.output_directory, f"FlexXray_Company_Summary_Page_{timestamp}.xlsx"
            )

        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Company Summary"

            # Define styles
            header_font = Font(bold=True, size=14)
            section_font = Font(bold=True, size=12)
            quote_font = Font(italic=True)
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font_white = Font(bold=True, color="FFFFFF")

            # Write title
            ws["A1"] = "FLEXXRAY COMPANY SUMMARY PAGE"
            ws["A1"].font = header_font
            ws.merge_cells("A1:D1")

            current_row = 3

            # Write key takeaways
            takeaways = summary_data.get("key_takeaways", [])
            if takeaways:
                ws[f"A{current_row}"] = "KEY TAKEAWAYS"
                ws[f"A{current_row}"].font = section_font
                current_row += 1

                for i, takeaway in enumerate(takeaways, 1):
                    ws[f"A{current_row}"] = f"{i}. {takeaway.get('insight', '')}"
                    ws[f"A{current_row}"].font = Font(bold=True)
                    current_row += 1

                    if takeaway.get("supporting_quotes"):
                        for quote in takeaway["supporting_quotes"][
                            :2
                        ]:  # Limit to 2 quotes
                            quote_cell = ws[f"B{current_row}"]
                            # Use formatted_text if available, otherwise fall back to text
                            quote_text = quote.get(
                                "formatted_text", quote.get("text", "")
                            )
                            quote_cell.value = f"• {quote_text}"
                            quote_cell.font = quote_font
                            quote_cell.alignment = Alignment(
                                wrap_text=False, vertical="top"
                            )
                            current_row += 1

                    current_row += 1

                current_row += 1

            # Write strengths
            strengths = summary_data.get("strengths", [])
            if strengths:
                ws[f"A{current_row}"] = "STRENGTHS"
                ws[f"A{current_row}"].font = section_font
                current_row += 1

                for i, strength in enumerate(strengths, 1):
                    ws[f"A{current_row}"] = f"{i}. {strength.get('insight', '')}"
                    ws[f"A{current_row}"].font = Font(bold=True)
                    current_row += 1

                    if strength.get("supporting_quotes"):
                        for quote in strength["supporting_quotes"][
                            :2
                        ]:  # Limit to 2 quotes
                            quote_cell = ws[f"B{current_row}"]
                            # Use formatted_text if available, otherwise fall back to text
                            quote_text = quote.get(
                                "formatted_text", quote.get("text", "")
                            )
                            quote_cell.value = f"• {quote_text}"
                            quote_cell.font = quote_font
                            quote_cell.alignment = Alignment(
                                wrap_text=False, vertical="top"
                            )
                            current_row += 1

                    current_row += 1

                current_row += 1

            # Write weaknesses
            weaknesses = summary_data.get("weaknesses", [])
            if weaknesses:
                ws[f"A{current_row}"] = "WEAKNESSES"
                ws[f"A{current_row}"].font = section_font
                current_row += 1

                for i, weakness in enumerate(weaknesses, 1):
                    ws[f"A{current_row}"] = f"{i}. {weakness.get('insight', '')}"
                    ws[f"A{current_row}"].font = Font(bold=True)
                    current_row += 1

                    if weakness.get("supporting_quotes"):
                        for quote in weakness["supporting_quotes"][
                            :2
                        ]:  # Limit to 2 quotes
                            quote_cell = ws[f"B{current_row}"]
                            # Use formatted_text if available, otherwise fall back to text
                            quote_text = quote.get(
                                "formatted_text", quote.get("text", "")
                            )
                            quote_cell.value = f"• {quote_text}"
                            quote_cell.font = quote_font
                            quote_cell.alignment = Alignment(
                                wrap_text=False, vertical="top"
                            )
                            current_row += 1

                    current_row += 1

            # Set specific column widths with special handling for quote column
            column_widths = {
                "A": 40,  # Main content column
                "B": 80,  # Quote column - Wide for full quote display
                "C": 30,  # Additional content columns
                "D": 30,
            }

            for col_letter, width in column_widths.items():
                if col_letter in ws.column_dimensions:
                    ws.column_dimensions[col_letter].width = width

            # Save workbook
            wb.save(output_file)
            wb.close()

            print(f"Company summary page exported to Excel: {output_file}")
            return output_file

        except (IOError, OSError) as e:
            print(f"File system error exporting company summary page to Excel: {e}")
            return ""
        except (ValueError, TypeError) as e:
            print(f"Data formatting error exporting company summary page to Excel: {e}")
            return ""
        except Exception as e:
            print(f"Unexpected error exporting company summary page to Excel: {e}")
            return ""

    def export_quotes_to_excel(
        self, quotes_data: List[Dict[str, Any]], output_file: str = None
    ) -> str:
        """Export quotes to Excel file with proper text wrapping for quote column."""
        if not EXCEL_AVAILABLE:
            print("openpyxl not available for Excel export")
            return ""

        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(
                self.output_directory, f"FlexXray_quotes_{timestamp}.xlsx"
            )

        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Quotes"

            # Define styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font_white = Font(bold=True, color="FFFFFF", size=12)
            quote_font = Font(size=11)

            # Define headers
            headers = [
                "Quote ID",
                "Quote Text",
                "Speaker",
                "Company/Title",
                "Transcript",
                "Sentiment",
                "Relevance Score",
                "Theme",
                "Date",
            ]

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # Write quote data
            for row, quote in enumerate(quotes_data, 2):
                # Quote ID
                ws.cell(row=row, column=1, value=quote.get("id", f"Q{row-1}"))

                # Quote Text - This is the main column that needs wrapping
                quote_cell = ws.cell(row=row, column=2, value=quote.get("text", ""))
                quote_cell.font = quote_font
                quote_cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Speaker
                speaker_info = quote.get("speaker_info", "")
                if isinstance(speaker_info, dict):
                    speaker_name = speaker_info.get("name", "")
                else:
                    speaker_name = str(speaker_info)
                ws.cell(row=row, column=3, value=speaker_name)

                # Company/Title
                if isinstance(speaker_info, dict):
                    company_title = speaker_info.get("company", "") or speaker_info.get(
                        "title", ""
                    )
                else:
                    company_title = ""
                ws.cell(row=row, column=4, value=company_title)

                # Transcript
                ws.cell(row=row, column=5, value=quote.get("transcript_name", ""))

                # Sentiment
                ws.cell(row=row, column=6, value=quote.get("sentiment", ""))

                # Relevance Score - Ensure numeric value
                relevance = quote.get("relevance_score", 0.0)
                if relevance is None or relevance == "" or relevance == "None":
                    relevance_value = 0.0
                else:
                    try:
                        relevance_value = float(relevance)
                    except (ValueError, TypeError):
                        relevance_value = 0.0
                ws.cell(row=row, column=7, value=relevance_value)

                # Theme
                ws.cell(row=row, column=8, value=quote.get("theme", ""))

                # Date
                date_value = quote.get("date", "")
                if date_value:
                    ws.cell(row=row, column=9, value=date_value)

            # Set column widths with special handling for quote column
            column_widths = {
                "A": 10,  # Quote ID
                "B": 80,  # Quote Text - Wide column for quotes
                "C": 20,  # Speaker
                "D": 25,  # Company/Title
                "E": 30,  # Transcript
                "F": 15,  # Sentiment
                "G": 15,  # Relevance Score
                "H": 20,  # Theme
                "I": 15,  # Date
            }

            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Set row heights for quote rows to accommodate wrapped text
            for row in range(2, len(quotes_data) + 2):
                ws.row_dimensions[row].height = 60  # Set minimum height for quote rows

            # Freeze the header row
            ws.freeze_panes = "A2"

            # Save workbook
            wb.save(output_file)
            wb.close()

            print(f"Quotes exported to Excel: {output_file}")
            return output_file

        except (IOError, OSError) as e:
            print(f"File system error exporting quotes to Excel: {e}")
            return ""
        except (ValueError, TypeError) as e:
            print(f"Data formatting error exporting quotes to Excel: {e}")
            return ""
        except Exception as e:
            print(f"Unexpected error exporting quotes to Excel: {e}")
            return ""

    def export_quote_analysis_to_excel(
        self, results: Dict[str, Any], output_file: str = None
    ) -> str:
        """Export quote analysis results to Excel with enhanced segmentation."""
        if not EXCEL_AVAILABLE:
            print("openpyxl not available for Excel export")
            return ""

        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(
                self.output_directory, f"FlexXray_Enhanced_Analysis_{timestamp}.xlsx"
            )

        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Define styles
            title_font = Font(bold=True, size=16, color="FFFFFF")
            header_font = Font(bold=True, size=14, color="FFFFFF")
            section_font = Font(bold=True, size=12, color="FFFFFF")
            question_font = Font(bold=True, size=11, color="000000")
            quote_font = Font(italic=True, size=10)
            metadata_font = Font(size=9, color="666666")
            
            title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            section_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            strength_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            weakness_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            takeaway_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            # ============================================================================
            # KEY TAKEAWAYS SHEET
            # ============================================================================
            ws_takeaways = wb.active
            if ws_takeaways is None:
                ws_takeaways = wb.create_sheet("Key Takeaways")
            else:
                ws_takeaways.title = "Key Takeaways"
            
            # Title
            ws_takeaways["A1"] = "KEY TAKEAWAYS & INSIGHTS"
            ws_takeaways["A1"].font = title_font
            ws_takeaways["A1"].fill = title_fill
            ws_takeaways.merge_cells("A1:F1")
            
            # Headers
            headers = ["Question", "Key Question", "Quote Text", "Speaker", "Transcript", "Relevance Score"]
            for col, header in enumerate(headers, 1):
                cell = ws_takeaways.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            current_row = 4
            
            # Process key takeaways
            key_takeaways = results.get("key_takeaways", [])
            for takeaway in key_takeaways:
                question = takeaway.get("question", "")
                question_key = takeaway.get("question_key", "")
                
                # Write question header
                question_cell = ws_takeaways[f"A{current_row}"]
                question_cell.value = question
                question_cell.font = question_font
                question_cell.fill = takeaway_fill
                question_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Question key
                ws_takeaways[f"B{current_row}"] = question_key
                ws_takeaways[f"B{current_row}"].fill = takeaway_fill
                
                # Merge question row
                ws_takeaways.merge_cells(f"A{current_row}:F{current_row}")
                current_row += 1
                
                # Write quotes for this takeaway
                quotes = takeaway.get("selected_quotes", [])
                for quote in quotes:
                    # Quote text
                    quote_cell = ws_takeaways[f"C{current_row}"]
                    quote_cell.value = quote.get("text", "")
                    quote_cell.font = quote_font
                    quote_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # Speaker info
                    metadata = quote.get("metadata", {})
                    speaker = metadata.get("transcript_name", "").split(" - ")[0] if metadata.get("transcript_name") else ""
                    ws_takeaways[f"D{current_row}"] = speaker
                    
                    # Transcript name
                    transcript = metadata.get("transcript_name", "")
                    ws_takeaways[f"E{current_row}"] = transcript
                    
                    # Relevance score
                    relevance = quote.get("relevance_score", 0)
                    ws_takeaways[f"F{current_row}"] = relevance
                    
                    current_row += 1
                
                current_row += 1  # Add space between takeaways
            
            # Set column widths
            column_widths = {"A": 50, "B": 20, "C": 80, "D": 25, "E": 40, "F": 15}
            for col_letter, width in column_widths.items():
                ws_takeaways.column_dimensions[col_letter].width = width
            
            # ============================================================================
            # STRENGTHS & WEAKNESSES SHEET
            # ============================================================================
            ws_swot = wb.create_sheet("Strengths & Weaknesses")
            
            # Title
            ws_swot["A1"] = "STRENGTHS & WEAKNESSES ANALYSIS"
            ws_swot["A1"].font = title_font
            ws_swot["A1"].fill = title_fill
            ws_swot.merge_cells("A1:F1")
            
            # Headers
            headers = ["Category", "Insight", "Quote Text", "Speaker", "Transcript", "Score"]
            for col, header in enumerate(headers, 1):
                cell = ws_swot.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            current_row = 4
            
            # Process strengths
            strengths = results.get("strengths", [])
            if strengths:
                # Strengths header
                strength_header = ws_swot[f"A{current_row}"]
                strength_header.value = "STRENGTHS"
                strength_header.font = section_font
                strength_header.fill = section_fill
                ws_swot.merge_cells(f"A{current_row}:F{current_row}")
                current_row += 1
                
                for strength in strengths:
                    # Insight
                    insight = strength.get("insight", strength.get("theme", ""))
                    ws_swot[f"B{current_row}"] = insight
                    ws_swot[f"B{current_row}"].font = question_font
                    ws_swot[f"B{current_row}"].fill = strength_fill
                    
                    # Category
                    ws_swot[f"A{current_row}"] = "Strength"
                    ws_swot[f"A{current_row}"].fill = strength_fill
                    
                    # Merge insight row
                    ws_swot.merge_cells(f"A{current_row}:F{current_row}")
                    current_row += 1
                    
                    # Supporting quotes
                    quotes = strength.get("supporting_quotes", strength.get("quotes", []))
                    for quote in quotes:
                        if isinstance(quote, dict):
                            quote_text = quote.get("formatted_text", quote.get("text", quote.get("quote", "")))
                            speaker = quote.get("speaker", quote.get("speaker_info", "Unknown"))
                            document = quote.get("document", quote.get("transcript_name", "Unknown"))
                            
                            ws_swot[f"C{current_row}"] = quote_text
                            ws_swot[f"C{current_row}"].font = quote_font
                            ws_swot[f"D{current_row}"] = speaker
                            ws_swot[f"E{current_row}"] = document
                            
                            current_row += 1
                    
                    current_row += 1
            
            # Process weaknesses
            weaknesses = results.get("weaknesses", [])
            if weaknesses:
                # Weaknesses header
                weakness_header = ws_swot[f"A{current_row}"]
                weakness_header.value = "WEAKNESSES"
                weakness_header.font = section_font
                weakness_header.fill = section_fill
                ws_swot.merge_cells(f"A{current_row}:F{current_row}")
                current_row += 1
                
                for weakness in weaknesses:
                    # Insight
                    insight = weakness.get("insight", weakness.get("theme", ""))
                    ws_swot[f"B{current_row}"] = insight
                    ws_swot[f"B{current_row}"].font = question_font
                    ws_swot[f"B{current_row}"].fill = weakness_fill
                    
                    # Category
                    ws_swot[f"A{current_row}"] = "Weakness"
                    ws_swot[f"A{current_row}"].fill = weakness_fill
                    
                    # Merge insight row
                    ws_swot.merge_cells(f"A{current_row}:F{current_row}")
                    current_row += 1
                    
                    # Supporting quotes
                    quotes = weakness.get("supporting_quotes", weakness.get("quotes", []))
                    for quote in quotes:
                        if isinstance(quote, dict):
                            quote_text = quote.get("formatted_text", quote.get("text", quote.get("quote", "")))
                            speaker = quote.get("speaker", quote.get("speaker_info", "Unknown"))
                            document = quote.get("document", quote.get("transcript_name", "Unknown"))
                            
                            ws_swot[f"C{current_row}"] = quote_text
                            ws_swot[f"C{current_row}"].font = quote_font
                            ws_swot[f"D{current_row}"] = speaker
                            ws_swot[f"E{current_row}"] = document
                            
                            current_row += 1
                    
                    current_row += 1
            
            # Set column widths for SWOT sheet
            column_widths = {"A": 15, "B": 50, "C": 80, "D": 25, "E": 40, "F": 15}
            for col_letter, width in column_widths.items():
                ws_swot.column_dimensions[col_letter].width = width
            
            # ============================================================================
            # QUOTES SUMMARY SHEET
            # ============================================================================
            ws_quotes = wb.create_sheet("Quotes Summary")
            
            # Title
            ws_quotes["A1"] = "QUOTES SUMMARY & METADATA"
            ws_quotes["A1"].font = title_font
            ws_quotes["A1"].fill = title_fill
            ws_quotes.merge_cells("A1:I1")
            
            # Headers
            headers = ["Quote ID", "Quote Text", "Speaker", "Company/Title", "Transcript", "Sentiment", "Relevance Score", "Theme", "Date"]
            for col, header in enumerate(headers, 1):
                cell = ws_quotes.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            current_row = 4
            
            # Collect all quotes from key takeaways
            all_quotes = []
            for takeaway in key_takeaways:
                quotes = takeaway.get("selected_quotes", [])
                for quote in quotes:
                    quote_data = {
                        "id": f"Q{len(all_quotes)+1:03d}",
                        "text": quote.get("text", ""),
                        "speaker": quote.get("metadata", {}).get("transcript_name", "").split(" - ")[0] if quote.get("metadata", {}).get("transcript_name") else "",
                        "company_title": quote.get("metadata", {}).get("transcript_name", "").split(" - ")[1] if quote.get("metadata", {}).get("transcript_name") and " - " in quote.get("metadata", {}).get("transcript_name", "") else "",
                        "transcript": quote.get("metadata", {}).get("transcript_name", ""),
                        "sentiment": "Positive" if quote.get("relevance_score", 0) >= 7 else "Neutral" if quote.get("relevance_score", 0) >= 5 else "Negative",
                        "relevance_score": quote.get("relevance_score", 0),
                        "theme": takeaway.get("question_key", ""),
                        "date": datetime.fromtimestamp(quote.get("metadata", {}).get("timestamp", 0)).strftime("%Y-%m-%d") if quote.get("metadata", {}).get("timestamp") else ""
                    }
                    all_quotes.append(quote_data)
            
            # Write quote data
            for quote_data in all_quotes:
                for col, (key, value) in enumerate(quote_data.items(), 1):
                    cell = ws_quotes.cell(row=current_row, column=col, value=value)
                    if key == "text":
                        cell.font = quote_font
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    elif key == "relevance_score":
                        cell.alignment = Alignment(horizontal="center")
                
                current_row += 1
            
            # Set column widths for quotes sheet
            column_widths = {"A": 10, "B": 80, "C": 20, "D": 25, "E": 30, "F": 15, "G": 15, "H": 20, "I": 15}
            for col_letter, width in column_widths.items():
                ws_quotes.column_dimensions[col_letter].width = width
            
            # Set row heights for quote rows
            for row in range(4, current_row):
                ws_quotes.row_dimensions[row].height = 60
            
            # Freeze panes
            ws_takeaways.freeze_panes = "A4"
            ws_swot.freeze_panes = "A4"
            ws_quotes.freeze_panes = "A4"
            
            # Save workbook
            wb.save(output_file)
            wb.close()
            
            print(f"Enhanced quote analysis exported to Excel: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"Error exporting enhanced quote analysis to Excel: {e}")
            return ""

    def _extract_section(self, text: str, section_header: str) -> str:
        """Extract a section from text based on header."""
        lines = text.split("\n")
        section_lines = []
        in_section = False

        for line in lines:
            if section_header.lower() in line.lower():
                in_section = True
                continue
            elif in_section and line.strip() and not line.startswith(" "):
                break
            elif in_section:
                section_lines.append(line)

        return "\n".join(section_lines).strip()

    def _extract_bullet_points(self, text: str, section_header: str) -> List[str]:
        """Extract bullet points from a section."""
        section_text = self._extract_section(text, section_header)
        if not section_text:
            return []

        lines = section_text.split("\n")
        bullet_points = []

        for line in lines:
            line = line.strip()
            if line.startswith(("•", "-", "*", "1.", "2.", "3.")):
                # Clean up the bullet point
                clean_line = re.sub(r"^[•\-*]\s*", "", line)
                clean_line = re.sub(r"^\d+\.\s*", "", clean_line)
                if clean_line:
                    bullet_points.append(clean_line)

        return bullet_points
